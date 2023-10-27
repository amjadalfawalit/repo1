from odoo.exceptions import AccessError, MissingError
from odoo.tools import image_process
import openpyxl
import xlsxwriter
import io
from odoo.osv.expression import AND, OR
from odoo.addons.portal.controllers.portal import CustomerPortal, pager as portal_pager
from odoo.tools import groupby as groupbyelem
from odoo.tools.translate import _
from odoo.http import request
from odoo import http
from markupsafe import Markup
from operator import itemgetter
import json
from odoo.addons.portal.controllers import portal
from datetime import datetime, timedelta
import base64

import logging
_logger = logging.getLogger(__name__)


class CustomerPortalInherit(CustomerPortal):

    def _prepare_home_portal_values(self, counters):
        values = super()._prepare_home_portal_values(counters)
        if 'location_count' in counters:
            location_count = request.env['stock.location'].search_count(self._get_location_portal_default_domain()) \
                if request.env['stock.location'].check_access_rights('read', raise_exception=False) else 0
            values['location_count'] = location_count

        if 'picking_count' in counters:
            picking_count = request.env['stock.picking'].search_count(self._get_stock_picking_portal_default_domain()) \
                if request.env['stock.picking'].check_access_rights('read', raise_exception=False) else 0
            values['picking_count'] = picking_count

        if 'product_count' in counters:
            product_count = request.env['product.template'].search_count(self._get_products_portal_default_domain()) \
                if request.env['product.template'].check_access_rights('read', raise_exception=False) else 0
            values['product_count'] = product_count

        return values

    def _get_stock_picking_search_domain(self, search_in, search):
        search_domain = []
        if search_in in ('all', 'name'):
            search_domain = OR([search_domain, [('name', 'ilike', search)]])
        return search_domain

    def _get_stock_picking_portal_default_domain(self):
        my_user = request.env.user
        return [
            ('partner_id', '=', my_user.partner_id.id),
        ]

    def _get_locations_search_domain(self, search_in, search):
        search_domain = []
        if search_in in ('all', 'name'):
            search_domain = OR([search_domain, [('name', 'ilike', search)]])
        return search_domain

    def _get_location_portal_default_domain(self):
        my_user = request.env.user
        return [
            ('current_partner_id', '=', my_user.partner_id.id),
        ]

    def _get_products_portal_default_domain(self):
        my_user = request.env.user
        return [
            ('member_id', '=', my_user.partner_id.id),
        ]

    def _get_products_search_domain(self, search_in, search):
        search_domain = []
        if search_in in ('all', 'name'):
            search_domain = OR([search_domain, [('name', 'ilike', search)]])
        return search_domain


    def render_success_template(self, success_message):
        return request.render('bmberry.excel_upload_success_template', {'success_message': success_message})

    def render_error_template(self, error_message):
        return request.render('bmberry.excel_upload_error_template', {'error_message': error_message})

    @http.route('/portal/upload/excel', type='http', auth="user", website=True, csrf=False)
    def upload_excel(self, **post):
        if 'excel_file' in request.httprequest.files:
            excel_file = request.httprequest.files['excel_file']
            if excel_file.filename:
                try:
                    # Load the uploaded Excel file
                    wb = openpyxl.load_workbook(excel_file)
                    my_user = request.env.user
                    errors = []

                    # Choose a specific worksheet to work with
                    # You can also specify a sheet by name: wb['Sheet1']
                    ws = wb.active
                    skip_first_row = True
                    row_index = 0
                    products = []
                    for row in ws.iter_rows(values_only=True):
                        # Access data from each row as a tuple
                        if skip_first_row:
                            skip_first_row = False
                            continue

                        row_index = row_index + 1
                        current_col_index = 0
                        # Perform data validation
                        try:
                            # Adjust column index as per your template
                            location_id = row[0]
                            location_name = row[1]
                            product_name = row[2]
                            category_name = row[3]
                            internal_reference = row[4]
                            sale_price = row[5]
                            barcode = row[6]
                            qty = row[7]
                            market_name = row[8]
                            scheduled_date = row[9]
                            location_id = int(row[current_col_index])
                            current_col_index += 1
                            location_name = str(row[current_col_index])
                            current_col_index += 1
                            product_name = str(row[current_col_index])
                            current_col_index += 1
                            category_name = str(row[current_col_index])
                            current_col_index += 1
                            internal_reference = str(row[current_col_index])
                            current_col_index += 1
                            sale_price = float(row[current_col_index])
                            current_col_index += 1
                            barcode = str(row[current_col_index])
                            current_col_index += 1
                            qty = float(row[current_col_index])
                            current_col_index += 1
                            market_name = str(row[current_col_index])
                            current_col_index += 1
                            scheduled_date = str(row[current_col_index])
                        except (ValueError, TypeError):
                            errors.append(
                                f"Row {row_index}, Column {current_col_index}: Data type validation failed.")
                            continue
                        # Add your custom validation logic here if needed
                        # ...

                    if errors:
                        error_message = '\n'.join(errors)

                        return self.render_error_template(error_message)

                    # Process the data from the worksheet
                    for row in ws.iter_rows(values_only=True):
                        if skip_first_row:
                            skip_first_row = False
                            continue
                        # Adjust column index as per your template
                        location_id = row[0]
                        location_name = row[1]
                        product_name = row[2]
                        category_name = row[3]
                        internal_reference = row[4]
                        sale_price = row[5]
                        barcode = row[6]
                        qty = row[7]
                        market_name = row[8]
                        scheduled_date = row[9]
                        _logger.info(row)
                        # Check if the product exists in the database based on product_name and barcode
                        product = request.env['product.product'].search([('name', '=', product_name), (
                            'member_barcode', '=', barcode), ('default_code', '=', internal_reference)], limit=1)
                        _logger.info(product)
                        if not product:
                            category = request.env['product.category'].search(
                                [('name', '=', category_name)], limit=1)
                            uom = request.env['uom.uom'].search(
                                [('name', 'ilike', 'Units')], limit=1)  # Adjust the search condition
                            if category and uom:
                                product = request.env['product.product'].create({
                                    'name': product_name,
                                    'member_barcode': barcode,
                                    'categ_id': category.id,
                                    'default_code': internal_reference,
                                    'detailed_type': 'product',
                                    'uom_id': uom.id,
                                    'uom_po_id': uom.id,
                                    'member_id': my_user.partner_id.id,
                                    'list_price': sale_price,
                                    # Add other product fields as needed
                                })

                        products.append(product)
                        # Create a receipt order with the specified details
                    if products:
                        _logger.info(
                            '==============location===================')
                        location_id = request.env['stock.location'].sudo().browse(
                            location_id)
                        _logger.info(
                            '==============location===================')

                        member_id = my_user.partner_id.id  # Get the member ID based on your
                        if location_id.current_partner_id.id != member_id:
                            return self.render_error_template(f"This Location Not Related To You Contact Support {location_id.name}")

                        _logger.info(
                            '==============location===================')

                        if location_id.warehouse_id.id == False:
                            return self.render_error_template(f"This Location Not Related To Market Contact Support {location_id.name}")

                        picking_type_id = request.env['stock.picking.type'].sudo().with_company(location_id.company_id.id).search(
                            [('code', '=', 'incoming'), ('warehouse_id', '=', location_id.warehouse_id.id), ('company_id', '=', location_id.warehouse_id.company_id.id)], limit=1)

                        _logger.info(
                            '==============location===================')

                        picking_vals = {
                            'partner_id': member_id,
                            'company_id': location_id.company_id.id,
                            'picking_type_id': picking_type_id.id,
                            'location_id': my_user.partner_id.property_stock_supplier.id,
                            'location_dest_id': location_id.id,
                            'owner_id': member_id,
                            'origin': f'Transfer from portal loading product from member :{str(my_user.partner_id.name)} to location : {location_id.name}',
                        }
                        _logger.info(
                            '==============picking===================')

                        _logger.info(picking_vals)

                        picking = request.env['stock.picking'].sudo().with_company(
                            location_id.company_id.id).create(picking_vals)

                        for product in products:
                            if product.id != False:
                                stock_move = request.env['stock.move'].sudo().with_company(location_id.company_id.id).create({
                                    'product_id': product.id,
                                    'product_uom': product.uom_id.id,
                                    'product_uom_qty': qty,
                                    'picking_id': picking.id,
                                    'location_id': my_user.partner_id.property_stock_supplier.id,
                                    'location_dest_id': location_id.id,
                                    'name': 'Item moved to member location',
                                })
                                stock_move_line = request.env['stock.move.line'].sudo().with_company(location_id.company_id.id).create({
                                    'product_id': product.id,
                                    'product_uom_id': product.uom_id.id,
                                    'qty_done': qty,
                                    'location_id': my_user.partner_id.property_stock_supplier.id,
                                    'location_dest_id': location_id.id,
                                    'move_id': stock_move.id,
                                    'picking_id': picking.id,
                                })
                        picking.action_confirm()
                        # picking._action_done()

                        _logger.info(product)

                    # You can add your custom data processing logic here

                    # Return a success message

                    return self.render_success_template(f"Excel file uploaded and processed successfully you can track your request by this refrence {picking.name}.")
                except Exception as e:
                    _logger.info(str(e))
                    return self.render_error_template(f"Error processing Excel file: {str(e)}")
        else:
            return self.render_error_template("No Excel file provided.")

    @http.route('/portal/download/excel/<int:locationId>', type='http', auth="user", website=True)
    def download_excel_template(self, locationId):
        # Create a workbook and add a worksheet
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output)
        worksheet = workbook.add_worksheet()
        location_id = request.env['stock.location'].browse(locationId)
        my_user = request.env.user

        # Fetch product categories from the database
        product_categories = request.env['product.category'].search([('is_public','=',True)])

        options = [category.name for category in product_categories]

        # Add content to the worksheet
        worksheet.write('A1', 'Location ID')
        worksheet.write('B1', 'Location Name')
        worksheet.write('C1', 'Product Name')
        worksheet.write('D1', 'Category')
        worksheet.write('E1', 'Internal Reference')
        worksheet.write('F1', 'Sale Price')
        worksheet.write('G1', 'Barcode')
        worksheet.write('H1', 'Qty')
        worksheet.write('I1', 'Market')
        worksheet.write('J1', 'Scheduled Date')

        worksheet.write('A2', locationId)
        worksheet.write('B2', location_id.name)
        worksheet.write('C2', '')  # Replace with actual product name
        worksheet.write('E2', '')  # Replace with actual internal reference
        worksheet.write('F2', 0.0)     # Replace with actual sale price
        worksheet.write('G2', '')       # Replace with actual barcode
        # Replace with actual quantity
        worksheet.write('H2', 1)
        # Replace with actual market
        worksheet.write('I2', location_id.warehouse_id.name)
        # Replace with actual scheduled date
        worksheet.write('J2', 'Scheduled Date Placeholder')

        # Create a dropdown list for the Category column
        worksheet.data_validation('D2:D1048576', {'validate': 'list',
                                                  'source': options,
                                                  'input_message': 'Select a category',
                                                  'error_message': 'Invalid category'})

        # Set column widths
        worksheet.set_column('A:A', 10)  # Location ID
        worksheet.set_column('B:B', 20)  # Location Name
        worksheet.set_column('C:C', 20)  # Product Name
        worksheet.set_column('D:D', 20)  # Category
        worksheet.set_column('E:E', 20)  # Internal Reference
        worksheet.set_column('F:F', 15)  # Sale Price
        worksheet.set_column('G:G', 15)  # Barcode
        worksheet.set_column('H:H', 10)  # Qty
        worksheet.set_column('I:I', 20)  # Market
        worksheet.set_column('J:J', 20)  # Scheduled Date

        # Close the workbook
        workbook.close()

        # Prepare the response with the Excel file
        response = request.make_response(
            output.getvalue(),
            headers=[
                ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition',
                 f'attachment; filename={location_id.name}_{location_id.warehouse_id.name}_{my_user.display_name}.xlsx')
            ]
        )
        response.mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        return response

    @http.route(['/my/locations', '/my/locations/page/<int:page>'], type='http', auth="user", website=True)
    def myLocation(self, page=1, date_begin=None, date_end=None, sortby='name', filterby='all', search=None, groupby='none', search_in='name', **kw):
        values = self._prepare_portal_layout_values()
        searchbar_sortings = {
            'date': {'label': _('Newest'), 'order': 'create_date desc'},
            'name': {'label': _('Subject'), 'order': 'name'},
            'reference': {'label': _('Reference'), 'order': 'id'},
        }
        searchbar_filters = {
            'all': {'label': _('All'), 'domain': []},
        }
        searchbar_inputs = {
            'content': {'input': 'name', 'label': _('Search In Name')},
            'id': {'input': 'id', 'label': _('Search in Reference')},
        }
        searchbar_groupby = {

            'none': {'input': 'none', 'label': _('None')},
        }

        domain = self._get_location_portal_default_domain()
        if search and search_in:
            domain = AND(
                [domain, self._get_locations_search_domain(search_in, search)])

        if not sortby:
            sortby = 'date'
        sort_order = searchbar_sortings[sortby]['order']

        # pager
        locations_count = request.env['stock.location'].search_count(domain)
        pager = portal_pager(
            url="/my/locations",
            url_args={'date_begin': date_begin, 'date_end': date_end, 'sortby': sortby, 'search_in': search_in,
                      'search': search, 'groupby': groupby, 'filterby': filterby},
            total=locations_count,
            page=page,
            step=self._items_per_page
        )

        locations = request.env['stock.location'].search(
            domain, limit=self._items_per_page, offset=pager['offset'], order=sort_order)

        values.update({
            'locations': locations,
            'date': date_begin,
            'page_name': 'locations',
            'default_url': '/my/locations',
            'pager': pager,
            'searchbar_sortings': searchbar_sortings,
            'searchbar_filters': searchbar_filters,
            'searchbar_inputs': searchbar_inputs,
            'searchbar_groupby': searchbar_groupby,
            'sortby': sortby,
            'groupby': groupby,
            'search_in': search_in,
            'search': search,
            'filterby': filterby,
        })
        return request.render("bmberry.portal_locations", values)

    @http.route(['/my/load_requests', '/my/load_requests/page/<int:page>'], type='http', auth="user", website=True)
    def myLoadRequests(self, page=1, date_begin=None, date_end=None, sortby='name', filterby='all', search=None, groupby='none', search_in='name', **kw):
        values = self._prepare_portal_layout_values()
        searchbar_sortings = {
            'date': {'label': _('Newest'), 'order': 'create_date desc'},
            'name': {'label': _('Subject'), 'order': 'name'},
            'reference': {'label': _('Reference'), 'order': 'id'},
        }
        searchbar_filters = {
            'all': {'label': _('All'), 'domain': []},
        }
        searchbar_inputs = {
            'content': {'input': 'name', 'label': _('Search In Name')},
            'id': {'input': 'id', 'label': _('Search in Reference')},
        }
        searchbar_groupby = {

            'none': {'input': 'none', 'label': _('None')},
        }

        domain = self._get_stock_picking_portal_default_domain()
        if search and search_in:
            domain = AND(
                [domain, self._get_stock_picking_search_domain(search_in, search)])

        if not sortby:
            sortby = 'date'
        sort_order = searchbar_sortings[sortby]['order']

        # pager
        picking_count = request.env['stock.picking'].search_count(domain)
        pager = portal_pager(
            url="/my/load_requests",
            url_args={'date_begin': date_begin, 'date_end': date_end, 'sortby': sortby, 'search_in': search_in,
                      'search': search, 'groupby': groupby, 'filterby': filterby},
            total=picking_count,
            page=page,
            step=self._items_per_page
        )

        pickings = request.env['stock.picking'].search(
            domain, limit=self._items_per_page, offset=pager['offset'], order=sort_order)

        values.update({
            'pickings': pickings,
            'date': date_begin,
            'page_name': 'stock_picking',
            'default_url': '/my/load_requests',
            'pager': pager,
            'searchbar_sortings': searchbar_sortings,
            'searchbar_filters': searchbar_filters,
            'searchbar_inputs': searchbar_inputs,
            'searchbar_groupby': searchbar_groupby,
            'sortby': sortby,
            'groupby': groupby,
            'search_in': search_in,
            'search': search,
            'filterby': filterby,
        })
        return request.render("bmberry.portal_stock_picking", values)

    def _transfer_get_page_view_values(self, order, access_token, **kwargs):
        #
        def resize_to_48(b64source):
            if not b64source:
                b64source = base64.b64encode(
                    request.env['ir.http']._placeholder())
            return image_process(b64source, size=(48, 48))

        values = {
            'order': order,
            'resize_to_48': resize_to_48,
            'report_type': 'html',
        }

        history = 'my_load_requests'
        return self._get_page_view_values(order, access_token, values, history, False, **kwargs)

    @http.route(['/my/stock/<int:order_id>'], type='http', auth="public", website=True)
    def portal_my_transfers_order(self, order_id=None, access_token=None, **kw):
        try:
            order_sudo = self._document_check_access(
                'stock.picking', order_id, access_token=access_token)
        except (AccessError, MissingError):
            return request.redirect('/my')

        report_type = kw.get('report_type')
        # if report_type in ('html', 'pdf', 'text'):
        # return self._show_report(model=order_sudo, report_type=report_type, report_ref='purchase.action_report_purchase_order', download=kw.get('download'))

        values = self._transfer_get_page_view_values(
            order_sudo, access_token, **kw)
        if order_sudo.company_id:
            values['res_company'] = order_sudo.company_id

        return request.render("bmberry.portal_my_load_request", values)

    @http.route(['/my/products', '/my/products/page/<int:page>'], type='http', auth="user", website=True)
    def myProducts(self, page=1, date_begin=None, date_end=None, sortby='name', filterby='all', search=None, groupby='none', search_in='name', **kw):
        values = self._prepare_portal_layout_values()
        searchbar_sortings = {
            'date': {'label': _('Newest'), 'order': 'create_date desc'},
            'name': {'label': _('Subject'), 'order': 'name'},
            'reference': {'label': _('Reference'), 'order': 'id'},
        }
        searchbar_filters = {
            'all': {'label': _('All'), 'domain': []},
        }
        searchbar_inputs = {
            'content': {'input': 'name', 'label': _('Search In Name')},
            'id': {'input': 'id', 'label': _('Search in Reference')},
        }
        searchbar_groupby = {

            'none': {'input': 'none', 'label': _('None')},
        }

        domain = self._get_products_portal_default_domain()
        if search and search_in:
            domain = AND(
                [domain, self._get_products_portal_search_domain(search_in, search)])

        if not sortby:
            sortby = 'date'
        sort_order = searchbar_sortings[sortby]['order']

        # pager
        products_count = request.env['product.template'].search_count(domain)
        pager = portal_pager(
            url="/my/products",
            url_args={'date_begin': date_begin, 'date_end': date_end, 'sortby': sortby, 'search_in': search_in,
                      'search': search, 'groupby': groupby, 'filterby': filterby},
            total=products_count,
            page=page,
            step=self._items_per_page
        )

        products = request.env['product.template'].search(
            domain, limit=self._items_per_page, offset=pager['offset'], order=sort_order)

        values.update({
            'products': products,
            'date': date_begin,
            'page_name': 'products',
            'default_url': '/my/products',
            'pager': pager,
            'searchbar_sortings': searchbar_sortings,
            'searchbar_filters': searchbar_filters,
            'searchbar_inputs': searchbar_inputs,
            'searchbar_groupby': searchbar_groupby,
            'sortby': sortby,
            'groupby': groupby,
            'search_in': search_in,
            'search': search,
            'filterby': filterby,
        })
        return request.render("bmberry.portal_products", values)
